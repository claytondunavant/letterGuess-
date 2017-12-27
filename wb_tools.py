import openpyxl
import os
import datetime
from collections import defaultdict

##########functions##########

#turns a list into a column on a workbook
def list_to_column (lists, workbook, sheet, column, start):
    print(list)
    workbookname = str(workbook) + '.xlsx'
    wb = openpyxl.load_workbook(workbookname)
    sheet = wb.get_sheet_by_name(str(sheet)) #assign sheet to variable

    for i in range(0,len(lists)): #loops through list and writes items in list to cells
        sheet[str(column) + str(start)].value = lists[i] #writes item in list to cell
        start = start + 1 #adds one to start to move down one cell

    wb.save(workbookname) #saves all the changes on workbook

#turns a column into a list: returns a list
def column_to_list (workbook, sheet, column, start):
    workbookname = str(workbook) + '.xlsx'
    wb = openpyxl.load_workbook(workbookname)
    sheet = wb.get_sheet_by_name(str(sheet))  # assign sheet to variable
    clist = [] #empty list for appending

    for i in range(0, sheet.max_row): #loops as many times as the max row in the column
        clist.append(sheet[str(column) + str(start)].value) #appends the value in the cell to the list
        start = start + 1 #adds one to start to move the cell down one

    return clist #returns the newly formed list

#creates new workbook
def new_wb (name):
    #og_file_path = os.getcwd() #gets original file path
    #wbs_file_path = os.getcwd() + '/workbooks/' #sets file path to the file path of the workbooks directory
    #os.chdir(wbs_file_path) #changes directory to the workbook directory
    wbname = str(name) + '.xlsx' #sets wbname to name
    wb = openpyxl.Workbook() #makes new workbook
    wb.save(wbname) #saves workbook
    #os.chdir(og_file_path) #goes back to orignal file path

#gets current date and time
def get_current_time ():
    now = datetime.datetime.now()  # sets up date and time
    return str(now.strftime("%Y-%m-%d %H %M %S"))

#takes the used letters and their frequencies and makes a wb
def letterfrequencies_to_wb (frequencies, usage, wbname):
    wbname = str(wbname) + '.xlsx' #sets name to appropiate file name
    wb = openpyxl.load_workbook(wbname) #opens workbook for use

    freq_keys = list(frequencies.keys()) #[1, 2, 3, 4, 5, etc.]

    for index in range(0, len(freq_keys)): #makes sheets
        sheet_name = freq_keys[index]
        wb.create_sheet(str(sheet_name))
        sheet = wb.get_sheet_by_name(str(sheet_name))

        letters = list(frequencies[sheet_name]) #['h', 'e', 'y', 't', 'r']
        start = 0

        for i in range(0, len(letters)):  # adds letters to sheets
            start = start + 1 #moves down rows
            sheet['A' + str(start)].value = letters[i] #adds letter
            sheet['B' + str(start)].value = usage[letters[i]] #adds letter's frequincy

        wb.save(wbname) #save workbook

##########output##########

