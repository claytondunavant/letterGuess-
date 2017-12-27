from random import randint
import openpyxl
from collections import defaultdict

##########functions##########

#returns random lowercase letter 
def random_letter():
    num = randint(0,25) #assigns num a random letter between and including 0 and 25
    letters = ['a','b','c', 'd','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z'] #list of all the letters in the alphabet
    return letters[num] #returns a random letter of the alphabet

#turns a string into a list: returns list
def string_to_list(string):
    output_list = []
    for i in range(0, len(string)):
        output_list.append(string[i])
    return output_list

#turns list into string: returns string
def list_to_string(list):
    return ''.join(list)

#used with sorted words excel document: returns phrase
def generate_phrase(length):
    print('generating phrase')
    letters = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 27, 28, 29, 31, 45] #all the sheets inside words
    phrase = [] #the output phrase
    sum = 0 #measures the amount of words in phrase before meeting the length requirement
    wb = openpyxl.load_workbook('words.xlsx') #loads sorted words workbook
    generated = False #sets generated to false because phrase hasn't been generated yet

    while generated == False:
        sheet_name = letters[randint(0, len(letters) - 1)] #chooses random sheet name out of list

        if sheet_name + sum == length: # If sheetname + sum = lentgh then the length is met
            sheet = wb.get_sheet_by_name(str(sheet_name)) #sets sheet to the actual sheet in worbook
            word = sheet['A' + str(randint(1,sheet.max_row))].value #sets word to a random cell in the sheet
            phrase.append(string_to_list(str(word))) #adds the value as a list to the phrase
            generated = True #the phrase has been generated so the loop ends
        if sheet_name + sum < length: # if sheetname + sum < lenght add the word to the phrase and continue with the loo[
            sum = sum + sheet_name #add the number of words to the sum of letters
            sheet = wb.get_sheet_by_name(str(sheet_name)) #sets sheet to the actual sheet in worbook
            word = sheet['A' + str(randint(1, sheet.max_row))].value #sets word to a random cell in the sheet
            phrase.append(string_to_list(str(word))) #adds the value as a list to the phrase

    return phrase

#updates answered if letter_guess is correct
def guess(letter_guess, answered, key, word, letter):
    if letter_guess == key[word][letter]:
        answered[word][letter] = letter_guess

'''
    for word in range(0,len(key)): #loop for how many words in phrase
        word = key[word] #a word is key[i]

        if lguess in word: #if the guess is in the word

            for letter in range(0,len(word)): #loop for how many letters in word
                letter = word[letter] #letter is word[g]

                if letter == lguess: #if letter == to guess
                    answered[word][letter] = lguess #update answered based on letter guesse
'''


#makes answered list full of zeros based on key: returns answered list
def key_to_answered(key):
    output_list = []

    for i in range(0, len(key)):
        word = []

        for g in range(0,len(key[i])):
            word.append(0)

        output_list.append(word)

    return output_list

#turns phrase into dict of frequencies: returns dic of frequencies and updates used
def phrase_to_frequencies(phrase, frequencies, usage):
    frequency_dict = frequencies
    used_dict = usage

    for word in phrase: #loop for each word

        for key in range(0,len(word)): #loop for each possible key
            key = key + 1 #key is upped by one to make logical sense to user
            word_skeleton = [] #sets up word skeleton for use

            for number in range(0,len(word)): #create list of all possible indexes in word
                word_skeleton.append(number)

            finished = False

            while finished == False: #creates instance of indexes of letters based on key
                last_index = 0
                letters = []

                for index in range(0,key): #for index in the confine of the max of the key value
                    word_skeleton[index] = word_skeleton[index] + 1 #updates index in word skeleton to
                    last_index = word_skeleton[index] #updates last_index to the last index used
                    letters.append(word[last_index - 1]) #adds charater to output list to be appended to output_dic

                if list_to_string(letters) not in used_dict: #if the letters have not been used yet, create an entry and set it to zero
                    used_dict[list_to_string(letters)] = 0
                    frequency_dict[key].append(list_to_string(letters))  # appends output_list to output_dict as a string

                used_dict[list_to_string(letters)] = used_dict[list_to_string(letters)] + 1 #add one to letter value because it has been used multiple times

                if last_index == len(word_skeleton): #if the last used index is the length of the word - end
                    finished = True

    return frequency_dict

#sorts used and frequencies dicts so they correspond to eachother
def sort_used_and_frequencies(frequencies, usage):
    keys = list(frequencies.keys())
    for i in range(0, len(keys)):
        frequencies[keys[i]].sort()

    used_keys = list(usage.keys())
    sorted_used_keys = []
    used_values = list(usage.values())
    sorted_used_dict = {}

    for i in range(0, len(used_keys)):
        sorted_used_keys.append(used_keys[i])

    sorted_used_keys.sort()

    for i in range(0, len(sorted_used_keys)):
        sorted_used_dict[sorted_used_keys[i]] = usage[sorted_used_keys[i]]

    used = sorted_used_dict

    return frequencies

def needed_letters_to_answer(phrase):
    needed_letters = []
    for word in range(0,len(phrase)):
        word = phrase[word]
        for letter in range(0,len(word)):
            letter = word[letter]
            if letter not in needed_letters:
                needed_letters.append(letter)
    return needed_letters



'''
def phrase_to_frequencies(phrase, used):
    frequency_dict = defaultdict(list) #sets up dictionary to be outputed
    used_dict = used

    for word in phrase: #loop for each word

        for key in range(0,len(word)): #loop for each possible key
            key = key + 1 #key is upped by one to make logical sense to user
            word_skeleton = [] #sets up word skeleton for use

            for number in range(0,len(word)): #create list of all possible indexes in word
                word_skeleton.append(number)

            finished = False

            while finished == False: #creates instance of indexes of letters based on key
                last_index = 0
                letters = []

                for index in range(0,key): #for index in the confine of the max of the key value
                    word_skeleton[index] = word_skeleton[index] + 1 #updates index in word skeleton to
                    last_index = word_skeleton[index] #updates last_index to the last index used
                    letters.append(word[last_index - 1]) #adds charater to output list to be appended to output_dic

                if list_to_string(letters) not in used_dict: #if the letters have not been used yet, create an entry and set it to zero
                    used_dict[list_to_string(letters)] = 0
                    frequency_dict[key].append(
                    list_to_string(letters))  # appends output_list to output_dict as a string

                used_dict[list_to_string(letters)] = used_dict[list_to_string(letters)] + 1 #add one to letter value because it has been used multiple times

                if last_index == len(word_skeleton): #if the last used index is the length of the word - end
                    finished = True

    print('frequency_dict: ' + str(frequency_dict))
    print('used_dict: ' + str(used_dict))

    return frequency_dict
'''
##########output##########

#phrase = [['h','e','y'],['t','h','e','r','e']]

