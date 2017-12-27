from collections import defaultdict
import phrase_tools
import wb_tools
import openpyxl
import time

#tell if words do share letters: returns true or false
def do_words_share_letters(letters, answered_letters):
    amount_of_same_letters = 0

    for index in range(len(answered_letters)):
        if letters[index] == answered_letters[index]:
            amount_of_same_letters = amount_of_same_letters + 1
        if amount_of_same_letters == len(answered_letters):
            return True

    return False

#returns most logical letter given the frequencies
def get_logical_letter(answered, frequencies, usage, used):

    random_letter = phrase_tools.random_letter()

    for word in range(len(answered)): #loop through words
        word = answered[word]
        answered_letters = []

        for letter in range(len(word)): #loop through letters in words
            letter = word[letter]

            if letter == 0: #if letter in word = 0
                frequency_key = len(answered_letters) + 1 #decides what key of frequency needs to be looked at

                if frequency_key not in list(frequencies.keys()): #if key does not exist in frequencies
                    print('not in frequencies')
                    return random_letter

                possible_letters = frequencies[frequency_key] #shows all possible letters

                if len(possible_letters) == 0: #if no possible letters
                    print('no possible letters in frequencies')
                    return random_letter

                ###if there are no answered letters###
                if frequency_key == 1:
                    print('no answered letters')
                    greatest_key = ''
                    greatest_value = 0

                    for i in range(0, len(possible_letters)):  # for all possible answers
                        current_letter = possible_letters[i]  # current letter out of possible answers

                        if current_letter not in used:
                            if usage[current_letter] > greatest_value:  # if the current letter usage is greater than the greatest value
                                greatest_key = current_letter  # greatest key = current letter
                                greatest_value = usage[current_letter]  # greatest value = usage of current letter

                    if greatest_key == '': #if all possible letters are in used
                        return random_letter

                    return greatest_key  # return greatest key


                ###if there are answered letters###
                sorted_possible_letters = []

                #sorts out not usable letters
                for i in range(0, len(possible_letters)):
                    pos_letter = possible_letters[i] #possible combo of letters
                    last_letter = str(phrase_tools.string_to_list(pos_letter)[len(pos_letter) - 1]) #gets last letter in pos_letter

                    if last_letter not in used: #if the last letter of pos_letter has not been used

                        if do_words_share_letters(pos_letter, answered_letters) == True: #if pos_letter and answered_letters dont share letters
                            sorted_possible_letters.append(last_letter)

                if len(sorted_possible_letters) == 0: #if there are not sorted possible letters
                    return random_letter


                greatest_key = ''
                greatest_value = 0

                #finds best letter to use
                for i in range(0, len(sorted_possible_letters)):  # for all possible answers
                    current_letter = sorted_possible_letters[i]# current letter out of possible answers

                    if current_letter not in used:
                        if usage[current_letter] > greatest_value:  # if the current letter usage is greater than the greatest value
                            greatest_key = current_letter  # greatest key = current letter
                            greatest_value = usage[current_letter]  # greatest value = usage of current letter

                if greatest_key == '':  # if all possible letters are in used
                    return random_letter

                return greatest_key  # return greatest key

            else:
                answered_letters.append(letter) #add letter to answered letters

def solve_random(length, min, wbname):

    wb_tools.new_wb(wbname) #creates new workbook
    wb = openpyxl.load_workbook(wbname + '.xlsx')
    wb.create_sheet('data')
    data = wb.get_sheet_by_name('data')
    wb.save(wbname + '.xlsx')

    rounds = 0

    min_met = False

    start = time.time()

    while min_met == False:
        rounds = rounds + 1 #updates to current round
        print('round: ' + str(rounds))
        key = phrase_tools.generate_phrase(length) #actual phrase
        print(key)
        needed_letters = phrase_tools.needed_letters_to_answer(key) #list of needed letters to answer phrase
        answered = phrase_tools.key_to_answered(key) #same length as phrase but full of zeros
        used = [] #used letters
        attempts = 0

        phrase_solved = False

        while phrase_solved == False:
            letter_guess = phrase_tools.random_letter() #sets letter_guess to random letter

            if letter_guess not in used: #if letter not used yet
                phrase_tools.guess(letter_guess, answered, key) #guesses the random letter
                used.append(letter_guess) #notes that the letter has been used
                attempts = attempts + 1 #adds one to attempts

                print(answered)

            if answered == key: #if answered has been solved
                data['A' + str(rounds)].value = str(rounds) #records the round
                data['B' + str(rounds)].value = str(attempts) + ':' + str(len(needed_letters)) #records attempts per length ratio
                data['C' + str(rounds)].value = str(float(attempts/len(needed_letters))) + ':1' #average number of attempts per an answer
                wb.save(wbname + '.xlsx')

                print('attempts per length ratio: ' + str(attempts) + ':' + str(len(needed_letters)))
                print('average attempts per a letter: ' + str(float(attempts/len(needed_letters))) + ':1')

                phrase_solved = True

                if float(attempts/len(needed_letters)) <= min:
                    min_met = True
        #phrase solved

    #min_met
    end = time.time()
    data['D' + str(rounds)].value = float(end - start)
    print('time: ' + str(float(end - start)))

def solve_logical(length, min, wbname):

    wb_tools.new_wb(wbname)  # creates new workbook
    wb = openpyxl.load_workbook(wbname + '.xlsx')
    wb.create_sheet('data')
    data = wb.get_sheet_by_name('data') #creates a sheet called data
    wb.save(wbname + '.xlsx') #saves wb

    rounds = 0
    frequencies = defaultdict(list)
    usage = {}

    min_met = False #min has not been met

    start = time.time() #start time

    while min_met == False: #while min has not been met
        rounds = rounds + 1  # updates to current round
        print('round: ' + str(rounds))
        key = phrase_tools.generate_phrase(length)  # actual phrase
        print(key)
        #needed_letters = phrase_tools.needed_letters_to_answer(key)  # list of needed letters to answer phrase
        answered = phrase_tools.key_to_answered(key)  # same length as phrase but full of zeros
        used = []  # used letters
        attempts = 0

        phrase_solved = False

        while phrase_solved == False:
            for w in range(0, len(key)):  # for each word
                key_word = key[w]  # word

                for l in range(0, len(key_word)):  # for each letter
                    used = []  # used resets each letter

                    while answered[w][l] == 0:  # will individual letter = 0
                        guess = get_logical_letter(answered, frequencies, usage, used)  # guess logical letter

                        if rounds == 1: #if its the first round guess random
                            phrase_tools.guess(phrase_tools.random_letter(), answered, key, w, l)
                            attempts = attempts + 1

                        if guess not in used:  # if the guess has not been used
                            phrase_tools.guess(guess, answered, key, w, l)  # guess
                            attempts = attempts + 1
                            used.append(guess)  # add guess to used
                            print(guess)
                            print(answered)

            data['A' + str(rounds)].value = str(rounds)  # records the round
            data['B' + str(rounds)].value = str(attempts) + ':' + str(
                length)  # records attempts per length ratio
            data['C' + str(rounds)].value = str(
                float(attempts / length))  # average number of attempts per an answer
            wb.save(wbname + '.xlsx')

            print('attempts per length ratio: ' + str(attempts) + ':' + str((length)))
            print('average attempts per a letter: ' + str(float(attempts / length)))

            phrase_solved = True

            if float(attempts / length) <= min: #victory condition
                end = time.time()
                data['D' + str(rounds)].value = float(end - start)
                print('time: ' + str(float(end - start)))
                wb.save(wbname + '.xlsx')
                min_met = True

        frequencies = phrase_tools.phrase_to_frequencies(key, frequencies, usage)
        frequencies = phrase_tools.sort_used_and_frequencies(frequencies, usage)
        wb_tools.letterfrequencies_to_wb(frequencies, usage, wbname)


##########output##########
